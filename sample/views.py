import io
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from .models import Brand, Buyer, Category, ChallengeImage, ChallengeIn, GG, Sample, StaffProfile
from .forms import BrandForm, BuyerForm, CategoryForm, ChallengeInForm, GGForm, SampleForm, StaffForm


def is_staff_or_admin(user):
    return user.is_staff or user.is_superuser


def is_superadmin(user):
    return user.is_superuser


def get_maker_user(user):
    """Return StaffProfile if the user is a maker account, else None."""
    try:
        return StaffProfile.objects.get(user=user)
    except StaffProfile.DoesNotExist:
        return None


def get_buyer_user(user):
    """Return Buyer if the user is a buyer account, else None."""
    try:
        return Buyer.objects.get(user=user)
    except Buyer.DoesNotExist:
        return None


# ---------- LOGIN ----------
def user_login(request):
    if request.method == 'POST':
        user = authenticate(
            request,
            username=request.POST['username'],
            password=request.POST['password'],
        )
        if user:
            login(request, user)
            return redirect('dashboard')
        return render(request, 'login.html', {'error': 'Invalid username or password. Please try again.'})
    return render(request, 'login.html')


@login_required
def user_logout(request):
    logout(request)
    return redirect('login')


# ---------- DASHBOARD ----------
@login_required
def dashboard(request):
    if request.user.is_superuser:
        sample_qs    = Sample.objects.all()
        total_buyers = Buyer.objects.count()
        total_makers = StaffProfile.objects.count()
    else:
        maker = get_maker_user(request.user)
        buyer = get_buyer_user(request.user)
        if maker:
            sample_qs = Sample.objects.filter(maker=maker).distinct()
        elif buyer:
            buyer_brands = buyer.brand.all()
            sample_qs = Sample.objects.filter(brand__in=buyer_brands).distinct()
        else:
            sample_qs = Sample.objects.none()
        total_buyers = 0
        total_makers = 0

    total_samples  = sample_qs.count()
    total_approved = sample_qs.filter(status='approved').count()
    total_pending  = sample_qs.filter(status='pending').count()
    total_draft    = sample_qs.filter(status='draft').count()
    total_rejected = sample_qs.filter(status='rejected').count()

    # Doughnut — status distribution
    status_chart = json.dumps({
        'labels': ['Approved', 'Pending', 'Draft', 'Rejected'],
        'data':   [total_approved, total_pending, total_draft, total_rejected],
        'colors': ['#22c55e', '#f59e0b', '#94a3b8', '#ef4444'],
    })

    # Bar — top buyers by sample count
    top_buyers = (
        sample_qs.filter(buyer__isnull=False)
        .values('buyer__buyer_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:8]
    )
    buyer_chart = json.dumps({
        'labels': [b['buyer__buyer_name'] for b in top_buyers],
        'data':   [b['count'] for b in top_buyers],
    })

    # Line — monthly submissions
    monthly = (
        sample_qs.filter(submission_date__isnull=False)
        .annotate(month=TruncMonth('submission_date'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    monthly_chart = json.dumps({
        'labels': [m['month'].strftime('%b %Y') for m in monthly],
        'data':   [m['count'] for m in monthly],
    })

    # Pie — samples by sample type
    by_type = (
        sample_qs.filter(sample_type__gt='')
        .values('sample_type')
        .annotate(count=Count('id'))
        .order_by('-count')[:6]
    )
    type_chart = json.dumps({
        'labels': [t['sample_type'] for t in by_type],
        'data':   [t['count'] for t in by_type],
    })

    brand_logos = Brand.objects.filter(logo__isnull=False).exclude(logo='')

    return render(request, 'dashboard.html', {
        'total_samples':  total_samples,
        'total_buyers':   total_buyers,
        'total_makers':   total_makers,
        'total_approved': total_approved,
        'total_pending':  total_pending,
        'total_draft':    total_draft,
        'total_rejected': total_rejected,
        'status_chart':   status_chart,
        'buyer_chart':    buyer_chart,
        'monthly_chart':  monthly_chart,
        'type_chart':     type_chart,
        'brand_logos':    brand_logos,
    })


# ---------- STAFF CRUD (SUPERADMIN ONLY) ----------
@login_required
@user_passes_test(is_superadmin)
def staff_list(request):
    staffs = StaffProfile.objects.select_related('user').all()
    return render(request, 'staff_list.html', {'staffs': staffs})


@login_required
@user_passes_test(is_superadmin)
def staff_detail(request, pk):
    staff   = get_object_or_404(StaffProfile.objects.select_related('user'), pk=pk)
    samples = Sample.objects.filter(maker=staff).select_related('buyer').prefetch_related('brand', 'gg', 'category').order_by('-submission_date')
    return render(request, 'staff_detail.html', {'staff': staff, 'samples': samples})


@login_required
@user_passes_test(is_superadmin)
def staff_create(request):
    form = StaffForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Staff member created successfully.')
        return redirect('staff_list')
    return render(request, 'form.html', {'form': form, 'title': 'Add Members'})


@login_required
@user_passes_test(is_superadmin)
def staff_update(request, pk):
    staff = get_object_or_404(StaffProfile, pk=pk)
    form = StaffForm(request.POST or None, instance=staff)
    if form.is_valid():
        form.save()
        messages.success(request, f'Staff member "{staff.user.username}" updated successfully.')
        return redirect('staff_list')
    return render(request, 'form.html', {'form': form, 'title': 'Edit Staff'})


@login_required
@user_passes_test(is_superadmin)
def staff_delete(request, pk):
    staff = get_object_or_404(StaffProfile, pk=pk)
    if request.method == 'POST':
        name = staff.user.username
        staff.user.delete()
        messages.success(request, f'Staff member "{name}" deleted successfully.')
        return redirect('staff_list')
    return render(request, 'confirm_delete.html', {'object': staff})


# ---------- BUYER CRUD ----------
@login_required
@user_passes_test(is_superadmin)
def buyer_list(request):
    buyers = Buyer.objects.select_related('user').prefetch_related('brand').all()
    return render(request, 'buyer_list.html', {'buyers': buyers})


@login_required
@user_passes_test(is_superadmin)
def buyer_detail(request, pk):
    buyer   = get_object_or_404(Buyer.objects.select_related('user').prefetch_related('brand'), pk=pk)
    samples = Sample.objects.filter(buyer=buyer).select_related('buyer').prefetch_related('brand', 'gg', 'category').order_by('-submission_date')
    return render(request, 'buyer_detail.html', {'buyer': buyer, 'samples': samples})


@login_required
@user_passes_test(is_superadmin)
def buyer_create(request):
    form = BuyerForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Buyer created successfully.')
        return redirect('buyer_list')
    return render(request, 'form.html', {'form': form, 'title': 'Add Buyer'})


@login_required
@user_passes_test(is_superadmin)
def buyer_update(request, pk):
    buyer = get_object_or_404(Buyer, pk=pk)
    form = BuyerForm(request.POST or None, instance=buyer)
    if form.is_valid():
        form.save()
        messages.success(request, f'Buyer "{buyer.buyer_name}" updated successfully.')
        return redirect('buyer_list')
    return render(request, 'form.html', {'form': form, 'title': 'Edit Buyer'})


@login_required
@user_passes_test(is_superadmin)
def buyer_delete(request, pk):
    buyer = get_object_or_404(Buyer, pk=pk)
    if request.method == 'POST':
        name = buyer.buyer_name
        buyer.user.delete()
        messages.success(request, f'Buyer "{name}" deleted successfully.')
        return redirect('buyer_list')
    return render(request, 'confirm_delete.html', {'object': buyer})


# ---------- CATEGORY CRUD ----------
@login_required
@user_passes_test(is_superadmin)
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'lookup_list.html', {
        'items': categories,
        'title': 'Categories',
        'create_url': 'category_create',
        'update_url': 'category_update',
        'delete_url': 'category_delete',
        'field': 'name',
    })


@login_required
@user_passes_test(is_superadmin)
def category_create(request):
    form = CategoryForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Category created successfully.')
        return redirect('category_list')
    return render(request, 'form.html', {'form': form, 'title': 'Add Category'})


@login_required
@user_passes_test(is_superadmin)
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    form = CategoryForm(request.POST or None, instance=category)
    if form.is_valid():
        form.save()
        messages.success(request, f'Category "{category.name}" updated successfully.')
        return redirect('category_list')
    return render(request, 'form.html', {'form': form, 'title': 'Edit Category'})


@login_required
@user_passes_test(is_superadmin)
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        name = category.name
        category.delete()
        messages.success(request, f'Category "{name}" deleted successfully.')
        return redirect('category_list')
    return render(request, 'confirm_delete.html', {'object': category})


# ---------- BRAND CRUD ----------
@login_required
@user_passes_test(is_superadmin)
def brand_list(request):
    brands = Brand.objects.annotate(style_count=Count('sample')).all()
    return render(request, 'brand_list.html', {'brands': brands})


@login_required
@user_passes_test(is_superadmin)
def brand_create(request):
    form = BrandForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Brand created successfully.')
        return redirect('brand_list')
    return render(request, 'form.html', {'form': form, 'title': 'Add Brand'})


@login_required
@user_passes_test(is_superadmin)
def brand_update(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    form = BrandForm(request.POST or None, request.FILES or None, instance=brand)
    if form.is_valid():
        form.save()
        messages.success(request, f'Brand "{brand.name}" updated successfully.')
        return redirect('brand_list')
    return render(request, 'form.html', {'form': form, 'title': 'Edit Brand'})


@login_required
@user_passes_test(is_superadmin)
def brand_delete(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        name = brand.name
        brand.delete()
        messages.success(request, f'Brand "{name}" deleted successfully.')
        return redirect('brand_list')
    return render(request, 'confirm_delete.html', {'object': brand})


# ---------- GG CRUD ----------
@login_required
@user_passes_test(is_superadmin)
def gg_list(request):
    ggs = GG.objects.all()
    return render(request, 'lookup_list.html', {
        'items': ggs,
        'title': 'GGs',
        'create_url': 'gg_create',
        'update_url': 'gg_update',
        'delete_url': 'gg_delete',
        'field': 'title',
    })


@login_required
@user_passes_test(is_superadmin)
def gg_create(request):
    form = GGForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'GG created successfully.')
        return redirect('gg_list')
    return render(request, 'form.html', {'form': form, 'title': 'Add GG'})


@login_required
@user_passes_test(is_superadmin)
def gg_update(request, pk):
    gg = get_object_or_404(GG, pk=pk)
    form = GGForm(request.POST or None, instance=gg)
    if form.is_valid():
        form.save()
        messages.success(request, f'GG "{gg.title}" updated successfully.')
        return redirect('gg_list')
    return render(request, 'form.html', {'form': form, 'title': 'Edit GG'})


@login_required
@user_passes_test(is_superadmin)
def gg_delete(request, pk):
    gg = get_object_or_404(GG, pk=pk)
    if request.method == 'POST':
        title = gg.title
        gg.delete()
        messages.success(request, f'GG "{title}" deleted successfully.')
        return redirect('gg_list')
    return render(request, 'confirm_delete.html', {'object': gg})


# ---------- CHALLENGE IN CRUD ----------
@login_required
@user_passes_test(is_superadmin)
def challengein_list(request):
    challenges = ChallengeIn.objects.all()
    return render(request, 'lookup_list.html', {
        'items': challenges,
        'title': 'Challenges In',
        'create_url': 'challengein_create',
        'update_url': 'challengein_update',
        'delete_url': 'challengein_delete',
        'field': 'title',
    })


@login_required
@user_passes_test(is_superadmin)
def challengein_create(request):
    form = ChallengeInForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Challenge In created successfully.')
        return redirect('challengein_list')
    return render(request, 'form.html', {'form': form, 'title': 'Add Challenge'})


@login_required
@user_passes_test(is_superadmin)
def challengein_update(request, pk):
    challenge = get_object_or_404(ChallengeIn, pk=pk)
    form = ChallengeInForm(request.POST or None, instance=challenge)
    if form.is_valid():
        form.save()
        messages.success(request, f'Challenge In "{challenge.title}" updated successfully.')
        return redirect('challengein_list')
    return render(request, 'form.html', {'form': form, 'title': 'Edit Challenge'})


@login_required
@user_passes_test(is_superadmin)
def challengein_delete(request, pk):
    challenge = get_object_or_404(ChallengeIn, pk=pk)
    if request.method == 'POST':
        title = challenge.title
        challenge.delete()
        messages.success(request, f'Challenge In "{title}" deleted successfully.')
        return redirect('challengein_list')
    return render(request, 'confirm_delete.html', {'object': challenge})


# ---------- helpers ----------
def _sample_queryset(request):
    """Return the base Sample queryset filtered by the current user's role."""
    base = Sample.objects.select_related('buyer').prefetch_related('gg', 'maker__user')
    if request.user.is_superuser:
        return base.all()
    maker = get_maker_user(request.user)
    if maker:
        return base.filter(maker=maker).distinct()
    buyer = get_buyer_user(request.user)
    if buyer:
        buyer_brands = buyer.brand.all()
        return base.filter(brand__in=buyer_brands).distinct()
    return Sample.objects.none()


def _apply_filters(qs, q, status):
    if status:
        qs = qs.filter(status=status)
    if q:
        qs = qs.filter(
            Q(style_number__icontains=q) |
            Q(buyer__buyer_name__icontains=q) |
            Q(sample_type__icontains=q) |
            Q(color__icontains=q) |
            Q(season__icontains=q)
        )
    return qs


# ---------- SAMPLE CRUD ----------
@login_required
def sample_list(request):
    q      = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    qs     = _apply_filters(_sample_queryset(request), q, status)
    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(request.GET.get('page'))
    return render(request, 'sample_list.html', {
        'samples':        page_obj,
        'page_obj':       page_obj,
        'total_count':    paginator.count,
        'total_buyers':   Buyer.objects.count(),
        'total_makers':   StaffProfile.objects.count(),
        'q':              q,
        'status':         status,
    })


@login_required
def sample_export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    q      = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    page   = request.GET.get('page', 1)
    qs     = _apply_filters(_sample_queryset(request), q, status)
    paginator = Paginator(qs, 10)
    samples   = list(paginator.get_page(page))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="samples.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=14, spaceAfter=10)

    header = ['#', 'Style No', 'Buyer', 'Sample Type', 'GG', 'Color', 'Season', 'Sub. Date', 'Status']
    data = [header]
    start = paginator.get_page(page).start_index()
    for i, s in enumerate(samples, start=start):
        gg_titles = ', '.join(g.title for g in s.gg.all()) or '—'
        data.append([
            str(i),
            s.style_number or '—',
            s.buyer.buyer_name if s.buyer else '—',
            s.sample_type or '—',
            gg_titles,
            s.color or '—',
            str(s.season) if s.season else '—',
            str(s.submission_date) if s.submission_date else '—',
            s.get_status_display(),
        ])

    col_widths = [1*cm, 3*cm, 3.5*cm, 3*cm, 3*cm, 2.5*cm, 2*cm, 2.8*cm, 2.2*cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0), 8),
        ('FONTSIZE',     (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
        ('ALIGN',        (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',      (0, 0), (-1, -1), 6),
    ]))

    doc.build([Paragraph('Sample List', title_style), Spacer(1, 0.3*cm), tbl])
    response.write(buffer.getvalue())
    return response


@login_required
def sample_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    q      = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    page   = request.GET.get('page', 1)
    qs     = _apply_filters(_sample_queryset(request), q, status)
    paginator = Paginator(qs, 10)
    page_obj  = paginator.get_page(page)
    samples   = list(page_obj)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Samples'

    header_fill   = PatternFill('solid', fgColor='1E293B')
    header_font   = Font(bold=True, color='FFFFFF', size=10)
    center        = Alignment(horizontal='center', vertical='center')
    left          = Alignment(horizontal='left', vertical='center')
    thin          = Side(style='thin', color='E2E8F0')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill      = PatternFill('solid', fgColor='F8FAFC')

    headers = ['#', 'Style No', 'Buyer', 'Sample Type', 'GG', 'Color', 'Season', 'Submission Date', 'Status']
    col_widths = [5, 18, 20, 18, 20, 14, 10, 18, 12]

    ws.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

    start = page_obj.start_index()
    for i, s in enumerate(samples, start=start):
        row = i - start + 2
        ws.row_dimensions[row].height = 18
        fill = alt_fill if (i % 2 == 0) else None
        gg_titles = ', '.join(g.title for g in s.gg.all()) or '—'
        values = [
            i,
            s.style_number or '—',
            s.buyer.buyer_name if s.buyer else '—',
            s.sample_type or '—',
            gg_titles,
            s.color or '—',
            s.season or '—',
            str(s.submission_date) if s.submission_date else '—',
            s.get_status_display(),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = center if col == 1 else left
            cell.border    = border
            if fill:
                cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="samples.xlsx"'
    wb.save(response)
    return response


@login_required
def sample_detail(request, pk):
    sample = get_object_or_404(Sample.objects.select_related('buyer').prefetch_related('maker__user'), pk=pk)
    if not (request.user.is_staff or request.user.is_superuser):
        try:
            buyer = Buyer.objects.get(user=request.user)
            if sample.buyer != buyer:
                messages.error(request, 'You do not have permission to view that sample.')
                return redirect('sample_list')
        except Buyer.DoesNotExist:
            return redirect('sample_list')
    return render(request, 'sample_detail.html', {'sample': sample})


@login_required
@user_passes_test(is_staff_or_admin)
def sample_create(request):
    form = SampleForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        sample = form.save()
        for img in request.FILES.getlist('challenge_images'):
            ChallengeImage.objects.create(sample=sample, image=img)
        messages.success(request, 'Sample created successfully.')
        return redirect('sample_detail', pk=sample.pk)
    return render(request, 'form.html', {'form': form, 'title': 'Add Sample'})


@login_required
@user_passes_test(is_staff_or_admin)
def sample_update(request, pk):
    sample = get_object_or_404(Sample, pk=pk)
    form = SampleForm(request.POST or None, request.FILES or None, instance=sample)
    if form.is_valid():
        form.save()
        # Delete removed challenge images
        for img_id in request.POST.getlist('delete_challenge_images'):
            ChallengeImage.objects.filter(pk=img_id, sample=sample).delete()
        # Add new challenge images
        for img in request.FILES.getlist('challenge_images'):
            ChallengeImage.objects.create(sample=sample, image=img)
        messages.success(request, 'Sample updated successfully.')
        return redirect('sample_detail', pk=sample.pk)
    challenge_images = sample.challenge_images.all()
    return render(request, 'form.html', {'form': form, 'title': 'Edit Sample', 'challenge_images': challenge_images})


@login_required
@user_passes_test(is_staff_or_admin)
def sample_delete(request, pk):
    sample = get_object_or_404(Sample, pk=pk)
    if request.method == 'POST':
        sample.delete()
        messages.success(request, 'Sample deleted successfully.')
        return redirect('sample_list')
    return render(request, 'confirm_delete.html', {'object': sample})
